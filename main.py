"""
Noxera Labs — RAG + General AI Chat
FastAPI + ChromaDB (persistent) + Claude
Multi-workspace: each tenant has its own collection + docs folder.
Documents are auto-ingested from ./docs/<workspace>/ on startup.
"""

import hashlib
import io
import json
import os
import re
import base64
import tempfile
from contextlib import asynccontextmanager
from pathlib import Path
from typing import AsyncIterator, Optional

import anthropic
import chromadb
import pdfplumber
from chromadb.utils.embedding_functions.onnx_mini_lm_l6_v2 import ONNXMiniLM_L6_V2
from openpyxl import Workbook, load_workbook
from fastapi import FastAPI, File, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ──────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────
DATA_DIR = Path("./data/chroma")
DATA_DIR.mkdir(parents=True, exist_ok=True)

TMP_DIR = Path(os.getenv("NOXERA_TMP_DIR", "./data/tmp"))
TMP_DIR.mkdir(parents=True, exist_ok=True)
TMP_DIR_ABS = str(TMP_DIR.resolve())
os.environ["TMPDIR"] = TMP_DIR_ABS
os.environ["TEMP"] = TMP_DIR_ABS
os.environ["TMP"] = TMP_DIR_ABS
tempfile.tempdir = TMP_DIR_ABS

ONNX_MODEL_DIR = Path(os.getenv("NOXERA_ONNX_MODEL_DIR", "./data/onnx_models/all-MiniLM-L6-v2"))
ONNX_MODEL_DIR.mkdir(parents=True, exist_ok=True)
ONNXMiniLM_L6_V2.DOWNLOAD_PATH = ONNX_MODEL_DIR.resolve()

DOCS_DIR = Path("./docs")
DOCS_DIR.mkdir(parents=True, exist_ok=True)

STATIC_DIR = Path("./static")
APP_VERSION = "5.0.0"
RAG_MIN_RELEVANCE = float(os.getenv("RAG_MIN_RELEVANCE", "0.18"))
MAX_RESPONSE_TOKENS = int(os.getenv("MAX_RESPONSE_TOKENS", "1600"))
MAX_RETRIEVAL_RESULTS = int(os.getenv("MAX_RETRIEVAL_RESULTS", "5"))
MAX_QUESTION_CHARS = int(os.getenv("MAX_QUESTION_CHARS", "4000"))
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(12 * 1024 * 1024)))
MAX_IMAGE_UPLOAD_BYTES = int(os.getenv("MAX_IMAGE_UPLOAD_BYTES", str(6 * 1024 * 1024)))
OCR_MODEL = os.getenv("ANTHROPIC_OCR_MODEL", "claude-haiku-4-5-20251001")
WEB_SEARCH_TOOL_TYPE = os.getenv("WEB_SEARCH_TOOL_TYPE", "web_search_20260209")
SUPPORTED_DOC_EXTS = {".pdf", ".txt", ".xlsx"}
IMAGE_MEDIA_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
}


def env_list(name: str, default: str = "") -> list[str]:
    value = os.getenv(name, default)
    return [item.strip() for item in value.split(",") if item.strip()]


ALLOWED_MODELS = set(
    env_list(
        "ALLOWED_MODELS",
        "claude-haiku-4-5-20251001,claude-sonnet-4-6,claude-opus-4-7",
    )
)

# ──────────────────────────────────────────────
# Workspaces (multi-tenant)
# ──────────────────────────────────────────────
NOXERA_PROMPT = """Du bist der KI-Assistent von Noxera Labs — einem unabhängigen Software- und KI-Studio aus Hamburg, Deutschland, gegründet von Noah Wilm und Raphael Ghazaryan.

Wenn Kontext-Abschnitte aus Dokumenten bereitgestellt werden:
- Nutze sie als primäre Grundlage deiner Antwort
- Verweise im Fließtext mit [1], [2] etc. auf die Quellen

Wenn kein Dokumentkontext vorhanden ist:
- Beantworte Fragen aus deinem allgemeinen Wissen

Wenn Nutzer fragen, wie sie eine Anfrage stellen, Noxera kontaktieren,
ein Projekt anfragen, einen Termin buchen oder "eine Anfrage" ohne weiteren
Kontext erwähnen:
- Verstehe das als Anfrage an Noxera Labs, nicht als allgemeine Erklärung,
  wie man dir als Chat-Assistent eine Frage stellt.
- Erkläre kurz den passenden Weg über die Noxera-Website:
  Allgemeine Anfragen: https://noxera-labs.de/anfragen
  Website- und App-Projekte: https://noxera-labs.de/onboarding
  KI-Automatisierung: https://noxera-labs.de/ki-anfrage
  Termin buchen: https://noxera-labs.de/termin
  Kontakt: https://noxera-labs.de/contact oder info@noxera-labs.de

Wenn die Antwort Noxera Labs, ein Projekt, eine Zusammenarbeit, Kontakt,
Website, Leistungen, Preise oder nächste Schritte betrifft:
- Füge am Ende knapp passende Links hinzu, z.B.
  Website: https://noxera-labs.de
  Kontakt: https://noxera-labs.de/contact
  Anfrage: https://noxera-labs.de/anfragen
  Termin: https://noxera-labs.de/termin
  E-Mail: info@noxera-labs.de
- Wähle nur Links, die zur Frage passen, und übertreibe es nicht.

Antworte stets präzise, hilfreich und in der Sprache der Frage."""

POPP_PROMPT = """Du bist der interne KI-Assistent von Popp Feinkost — einem traditionsreichen Hersteller hochwertiger Feinkostsalate, Brotaufstriche und Convenience-Produkte aus Norddeutschland.

Du hilfst Mitarbeitenden in Produktion, Qualitätssicherung, Logistik und Verwaltung dabei, Antworten aus internen Dokumenten zu finden:
- Produktionshandbücher und Rezepturen
- Maschinenhandbücher, Bedienungsanleitungen und Störungsleitfäden
- Wartungspläne, Rüstvorgaben, CIP-Reinigung und Ersatzteilhinweise
- HACCP-Pläne und Hygienevorschriften
- Allergeninformationen und Produkt-Spezifikationen
- Lieferanten- und Rohstofflisten
- Schichtpläne, SOPs und Excel-Auswertungen

Wenn Kontext-Abschnitte aus Dokumenten bereitgestellt werden:
- Nutze sie als primäre und einzige Grundlage deiner Antwort
- Verweise im Fließtext mit [1], [2] etc. auf die Quellen
- Bei Tabellen-Daten (Excel): Antworte präzise mit konkreten Werten, Mengen, Temperaturen, Zeiten und Verantwortlichkeiten
- Wenn eine Information nicht im Dokumentkontext steht, sag das klar — erfinde keine Werte oder Verfahren

Wenn die Frage über die hinterlegten Dokumente hinausgeht (z. B. tagesaktuelle Schichten, individuelle Personalfragen, Preise):
- Weise freundlich darauf hin, dass die Information nicht in der Wissensbasis liegt
- Empfiehl die zuständige Abteilung: Qualitätssicherung, Produktionsleitung, Technik/Instandhaltung, Personalabteilung oder Einkauf

Antworte stets präzise, sachlich, in fachlich korrekter Sprache und immer in der Sprache der Frage — bevorzugt Deutsch."""


WORKSPACES: dict = {
    "noxera": {
        "id": "noxera",
        "name": "Noxera Labs",
        "tagline": "RAG-Demo von Noxera Labs",
        "description": "Finde Antworten in Firmendokumenten, ohne Ordner, PDFs und alte Dateien manuell zu durchsuchen. Der Assistent nutzt eure Wissensbasis und zeigt Quellen direkt mit an.",
        "accent": "#00CFFF",
        "logo_url": "https://noxera-labs.de/logo.svg",
        "allow_web_search": True,
        "system_prompt": NOXERA_PROMPT,
        "chips": [
            {"label": "Wissensbasis", "q": "Welche Dokumente sind in der Wissensbasis und was steht darin?"},
            {"label": "Kurz-Pitch", "q": "Fasse Noxera Labs so zusammen, dass ein Geschäftsführer sofort versteht, warum das relevant ist."},
            {"label": "Prozessanalyse", "q": "Analysiere einen typischen Firmenprozess und zeige, wie ein RAG-System Dokumentensuche und Support verbessern kann."},
            {"label": "KI-Lösungen", "q": "Welche KI-Lösungen bietet Noxera Labs an und welche wären für ein mittelständisches Unternehmen sinnvoll?"},
            {"label": "Anfrage", "q": "Erstelle eine kurze Projektanfrage an Noxera Labs mit Ziel, Datenquellen, Integrationen und nächsten Schritten."},
        ],
    },
    "popp": {
        "id": "popp",
        "name": "Popp Feinkost",
        "tagline": "Interner Wissensassistent — Popp Feinkost",
        "description": "Antworten aus Produktionshandbüchern, Maschinenanleitungen, Wartungsplänen, Rezepturen, HACCP-Plänen und Excel-Tabellen — direkt aus eurer internen Wissensbasis. Mit Quellenangabe.",
        "accent": "#E8472D",
        "logo_url": None,
        "allow_web_search": False,
        "system_prompt": POPP_PROMPT,
        "chips": [
            {"label": "Heringssalat-Prozess", "q": "Beschreibe Schritt für Schritt den Produktionsprozess für Heringssalat inklusive Temperaturen und Hygieneanforderungen."},
            {"label": "Rezeptur", "q": "Wie ist die exakte Rezeptur für Heringssalat — Mengen, Lieferanten, Allergene?"},
            {"label": "Maschinenstart", "q": "Wie starte ich die Abfüllanlage AF-2400 auf Linie 3 korrekt und welche Prüfungen sind vor Produktionsbeginn nötig?"},
            {"label": "Störung beheben", "q": "Was soll ich tun, wenn die Abfüllanlage AF-2400 den Fehler E-17 oder ungleichmäßige Füllgewichte meldet?"},
            {"label": "Wartung Linie 3", "q": "Welche täglichen und wöchentlichen Wartungsarbeiten sind an Linie 3 vorgeschrieben?"},
            {"label": "HACCP-CCPs", "q": "Welche kritischen Kontrollpunkte (CCPs) gibt es laut HACCP-Plan und was sind die Grenzwerte?"},
            {"label": "Allergene", "q": "Welche Allergene sind in unseren Brotaufstrichen enthalten und welche Kreuzkontaminationsrisiken gibt es?"},
            {"label": "CIP-Reinigung", "q": "Wie läuft die CIP-Reinigung am Mischer M-7 ab und welche Freigabewerte müssen erreicht werden?"},
        ],
    },
}
DEFAULT_WORKSPACE = os.getenv("DEFAULT_WORKSPACE", "popp")
if DEFAULT_WORKSPACE not in WORKSPACES:
    DEFAULT_WORKSPACE = "noxera"


def workspace_id_or_default(ws: Optional[str]) -> str:
    return ws if ws and ws in WORKSPACES else DEFAULT_WORKSPACE


def web_search_allowed(ws_id: str) -> bool:
    return bool(WORKSPACES.get(ws_id, {}).get("allow_web_search", False))


def docs_dir_for(ws_id: str) -> Path:
    p = DOCS_DIR / ws_id
    p.mkdir(parents=True, exist_ok=True)
    return p


# ──────────────────────────────────────────────
# ChromaDB (collection per workspace, lazy)
# ──────────────────────────────────────────────
chroma_client = chromadb.PersistentClient(path=str(DATA_DIR))
embedding_fn = ONNXMiniLM_L6_V2()
_collections: dict = {}


def collection_for(ws_id: str):
    ws_id = workspace_id_or_default(ws_id)
    if ws_id not in _collections:
        _collections[ws_id] = chroma_client.get_or_create_collection(
            name=f"ws_{ws_id}_documents",
            embedding_function=embedding_fn,
            metadata={"hnsw:space": "cosine"},
        )
    return _collections[ws_id]


# ──────────────────────────────────────────────
# Anthropic
# ──────────────────────────────────────────────
anthropic_client = anthropic.AsyncAnthropic()
DEFAULT_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-6")


def system_prompt_for(ws_id: str) -> str:
    return WORKSPACES[workspace_id_or_default(ws_id)]["system_prompt"]


# ──────────────────────────────────────────────
# Text helpers
# ──────────────────────────────────────────────
def chunk_text(text: str, chunk_size: int = 400, overlap: int = 60) -> list[str]:
    words = text.split()
    chunks = []
    step = max(1, chunk_size - overlap)
    for i in range(0, len(words), step):
        chunk = " ".join(words[i : i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
    return chunks


def extract_pdf_text(content: bytes) -> str:
    text = ""
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text.strip()


def extract_excel_sheets(content: bytes) -> list[tuple[str, str]]:
    """Convert .xlsx sheets to labeled text sections for embedding."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheets: list[tuple[str, str]] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [
                    ("" if c is None else str(c).strip().replace("\n", " "))
                    for c in row
                ]
                if any(cells):
                    rows_text.append(" | ".join(cells))
            if rows_text:
                sheets.append((sheet_name, f"=== Tabelle: {sheet_name} ===\n" + "\n".join(rows_text)))
    finally:
        wb.close()
    return sheets


def extract_excel_text(content: bytes) -> str:
    """Convert .xlsx to a text representation suitable for embedding."""
    return "\n\n".join(text for _, text in extract_excel_sheets(content)).strip()


def file_hash(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()[:16]


def safe_filename(filename: str) -> str:
    name = Path(filename or "upload").name
    stem = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(name).stem).strip(".-") or "upload"
    suffix = Path(name).suffix.lower()
    return f"{stem}{suffix}"


def _all_docs(ws_id: str) -> list[Path]:
    base = docs_dir_for(ws_id)
    files: list[Path] = []
    for ext in SUPPORTED_DOC_EXTS:
        files.extend(base.rglob(f"*{ext}"))
    return sorted(files)


def docs_summary(ws_id: str) -> dict:
    files = _all_docs(ws_id)
    return {
        "files": len(files),
        "pdfs": sum(1 for f in files if f.suffix.lower() == ".pdf"),
        "txts": sum(1 for f in files if f.suffix.lower() == ".txt"),
        "xlsx": sum(1 for f in files if f.suffix.lower() == ".xlsx"),
    }


def docs_payload(ws_id: str) -> dict:
    base = docs_dir_for(ws_id)
    files = _all_docs(ws_id)
    updated_at = max((f.stat().st_mtime for f in files), default=None)
    return {
        "workspace": ws_id,
        "summary": docs_summary(ws_id),
        "chunks_indexed": collection_for(ws_id).count(),
        "updated_at": updated_at,
        "files": [
            {
                "filename": f.name,
                "type": f.suffix.lower().lstrip("."),
                "size_bytes": f.stat().st_size,
                "source_path": str(f.relative_to(base)),
            }
            for f in files
        ],
    }


def _index_text(
    ws_id: str,
    filename: str,
    text: str,
    source_path: str = "",
    sheet_name: Optional[str] = None,
) -> tuple[str, int]:
    """Index extracted text into the workspace collection."""
    coll = collection_for(ws_id)
    fhash = file_hash(text.encode())
    stem = Path(filename).stem
    doc_id = f"{stem}__{fhash}"

    existing = coll.get(where={"doc_id": doc_id}, limit=1)
    if existing.get("ids"):
        return ("skipped", 0)

    chunks = chunk_text(text)
    if not chunks:
        return ("too_short", 0)

    chunk_ids = [f"{doc_id}_{i}" for i in range(len(chunks))]
    metadatas = [
        {
            "doc_id": doc_id,
            "filename": filename,
            "chunk_index": i,
            "source_path": source_path or filename,
            "workspace": ws_id,
            "sheet_name": sheet_name or "",
        }
        for i in range(len(chunks))
    ]
    coll.add(ids=chunk_ids, documents=chunks, metadatas=metadatas)
    return ("indexed", len(chunks))


def _index_excel_bytes(ws_id: str, filename: str, content: bytes, source_path: str = "") -> tuple[str, int, set[str]]:
    sheets = extract_excel_sheets(content)
    if not sheets:
        return ("empty", 0, set())

    statuses: list[str] = []
    total_chunks = 0
    doc_ids: set[str] = set()
    stem = Path(filename).stem
    for sheet_name, text in sheets:
        doc_ids.add(f"{stem}__{file_hash(text.encode())}")
        status, n = _index_text(
            ws_id,
            filename,
            text,
            f"{source_path or filename} / {sheet_name}",
            sheet_name=sheet_name,
        )
        statuses.append(status)
        total_chunks += n

    if any(s == "indexed" for s in statuses):
        return ("indexed", total_chunks, doc_ids)
    if all(s == "skipped" for s in statuses):
        return ("skipped", total_chunks, doc_ids)
    return (statuses[0], total_chunks, doc_ids)


def _read_file_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf_text(path.read_bytes())
    if suffix == ".xlsx":
        return extract_excel_text(path.read_bytes())
    return path.read_text(encoding="utf-8", errors="ignore").strip()


def sync_docs_folder(ws_id: str):
    coll = collection_for(ws_id)
    base = docs_dir_for(ws_id)
    files = _all_docs(ws_id)
    seen_doc_ids: set[str] = set()
    summary = {"indexed": 0, "skipped": 0, "empty": 0, "too_short": 0, "removed": 0, "errors": 0}

    for f in files:
        try:
            source_path = str(f.relative_to(base))
            if f.suffix.lower() == ".xlsx":
                status, _, doc_ids = _index_excel_bytes(ws_id, f.name, f.read_bytes(), source_path)
                seen_doc_ids.update(doc_ids)
            else:
                text = _read_file_text(f)
                if not text:
                    status = "empty"
                else:
                    doc_id = f"{f.stem}__{file_hash(text.encode())}"
                    seen_doc_ids.add(doc_id)
                    status, _ = _index_text(ws_id, f.name, text, source_path)
            summary[status] = summary.get(status, 0) + 1
        except Exception as e:
            summary["errors"] += 1
            print(f"[ingest:{ws_id}] {f.name}: error {e}")

    try:
        if summary["errors"] == 0:
            existing = coll.get()
            stale_ids: set[str] = set()
            for meta in existing.get("metadatas", []) or []:
                if not meta:
                    continue
                d = meta.get("doc_id")
                if d and d not in seen_doc_ids:
                    stale_ids.add(d)
            for d in stale_ids:
                entries = coll.get(where={"doc_id": d})
                if entries.get("ids"):
                    coll.delete(ids=entries["ids"])
                    summary["removed"] += 1
        else:
            print(f"[ingest:{ws_id}] cleanup skipped because indexing had errors")
    except Exception as e:
        print(f"[ingest:{ws_id}] cleanup warning: {e}")

    print(f"[ingest:{ws_id}] sync complete: {summary} (total chunks: {coll.count()})")
    return summary


# ──────────────────────────────────────────────
# Demo data seeding
# ──────────────────────────────────────────────
def _migrate_legacy_docs():
    """Move legacy ./docs/<file> → ./docs/noxera/<file> (one-time)."""
    target = docs_dir_for("noxera")
    for p in DOCS_DIR.iterdir():
        if p.is_file() and p.suffix.lower() in SUPPORTED_DOC_EXTS:
            dest = target / p.name
            if not dest.exists():
                p.rename(dest)
                print(f"[migrate] {p.name} → noxera/")


def _seed_popp_demo():
    """Auto-create example documents for the Popp Feinkost workspace."""
    popp_dir = docs_dir_for("popp")
    has_docs = any(p for p in popp_dir.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_DOC_EXTS)
    if not has_docs:
        print("[seed] generating Popp Feinkost demo documents…")
        _seed_popp_text(popp_dir)
        _seed_popp_excel(popp_dir)
    _seed_popp_machine_text(popp_dir)


def _write_demo_doc(path: Path, content: str):
    if not path.exists():
        path.write_text(content.strip() + "\n", encoding="utf-8")


def _seed_popp_text(popp_dir: Path):
    (popp_dir / "produktionshandbuch_heringssalat.txt").write_text(
        """PRODUKTIONSHANDBUCH HERINGSSALAT NACH HAUSFRAUENART
Popp Feinkost — Werk Kaltenkirchen / Linie 3
Dokumenten-Nr.: PH-HSA-2026-01
Freigegeben durch: Produktionsleitung & QS
Letzte Revision: 2026-02-12

1. PRODUKTBESCHREIBUNG
Heringssalat nach Hausfrauenart, gewürfelte Heringsfilets in Joghurt-Mayonnaise-Dressing
mit Roter Bete, Apfel, Gewürzgurken und Zwiebeln. Mindesthaltbarkeit: 28 Tage gekühlt
bei +4 °C in luftdichten Bechern (250 g und 500 g).

2. ZUTATEN UND CHARGEN
Hauptkomponenten:
- Matjes-Heringsfilets, entgrätet (38 % Anteil) — Lieferant Lemvig Fisheries A/S
- Rote Bete in Würfeln (16 %) — Lieferant Nordmark Gemüse GmbH
- Gewürzgurken (10 %) — eigene Einlegung
- Äpfel Boskoop (8 %) — saisonal aus Altes Land
- Mayonnaise 80 % (15 %) — eigene Herstellung Linie 5
- Joghurt 3,5 % (10 %) — Molkerei Holstein
- Zwiebeln gewürfelt (3 %)

3. PRODUKTIONSSCHRITTE
3.1 Vorbereitung (Schritt 1)
- Heringsfilets aus Salzlake entnehmen, 90 Minuten in Frischwasser bei +4 °C wässern
- pH-Wert nach Wässerung: Ziel 5,2 ± 0,2
- Salzgehalt nach Wässerung: 2,8 % ± 0,3 %
- Filets in 12 mm Würfel schneiden (Maschine Treif Puma 700)

3.2 Gemüsezubereitung (Schritt 2)
- Rote Bete blanchieren bei 92 °C für 8 Minuten, dann auf +4 °C abkühlen
- Apfel in Stücke schneiden und in 0,3 % Citratlösung tauchen (Bräunungsschutz)
- Zwiebeln 30 Sekunden in 80 °C heißem Wasser blanchieren

3.3 Dressing-Herstellung (Schritt 3)
- Mayonnaise und Joghurt im Verhältnis 60:40 mischen
- Salz, Zucker, Senf und Gewürze nach Rezeptur PA-DRS-014 zugeben
- Mischtemperatur: konstant unter +6 °C
- Endkontrolle pH-Wert: 4,1 ± 0,1

3.4 Mischen und Abfüllen (Schritt 4)
- Komponenten in Mischer M-7 vorsichtig untereinanderheben (max. 8 U/min)
- Mischzeit: 4 Minuten — länger zerstört die Heringswürfel
- Sofort in vorgekühlte Becher abfüllen, Siegelfolie aufbringen
- Becher müssen Siegeltest bestehen (Drucktest 0,3 bar / 5 Sekunden)

4. KRITISCHE KONTROLLPUNKTE
- CCP-1: Wässerung der Heringe — Salzgehalt
- CCP-2: Dressing-Temperatur — max. +6 °C
- CCP-3: Abfülltemperatur — Produkt darf nicht über +7 °C steigen
- CCP-4: Siegelnaht — 100 % Kontrolle durch Inline-Vision-System

5. QUALITÄTSKONTROLLE NACH PRODUKTION
- Sensorische Prüfung pro Charge: Aussehen, Geruch, Geschmack, Konsistenz
- Mikrobiologie-Stichprobe: Listeria, Salmonellen (Ergebnis nach 48 h)
- Gewichtskontrolle: ± 2 g pro Becher (250 g)

6. DOKUMENTATION
Jeder Produktionslauf wird im SAP-System unter Auftragsnummer und Charge
dokumentiert. Abweichungen werden im NCR-System erfasst und an die QS-Leitung
gemeldet.

ANSPRECHPARTNER
Produktionsleitung Linie 3: Stefan Brügge — Durchwahl 1432
Qualitätssicherung: Karin Petersen — Durchwahl 1810
Schichtleitung Frühschicht: Marek Wojciechowski — Durchwahl 1438
""",
        encoding="utf-8",
    )

    (popp_dir / "hygieneplan_produktion.txt").write_text(
        """HYGIENEPLAN PRODUKTION
Popp Feinkost — Standorte Kaltenkirchen und Bremerhaven
Dokumenten-Nr.: HYG-PRO-2026-04
Freigegeben durch: QS-Leitung
Letzte Revision: 2026-01-08

1. GELTUNGSBEREICH
Dieser Hygieneplan gilt für alle Produktionsbereiche, Lager- und Versandflächen,
inklusive Fischverarbeitung (Linie 1-3), Salatherstellung (Linie 4-6),
Brotaufstrich-Linie (Linie 7) und das Hochregallager Kaltenkirchen.

2. PERSONALHYGIENE
2.1 Berufskleidung
- Saubere Schutzkleidung wird täglich gewechselt
- Kopfbedeckung in allen Produktionsräumen verpflichtend
- Bartträger müssen Bartnetz tragen
- Schmuck (außer schlichter Ehering) ist verboten
- Fingernägel kurz, ohne Lack, keine künstlichen Nägel

2.2 Händehygiene
- Hände waschen vor Arbeitsbeginn, nach Pausen, nach Toilettengang
- Desinfektion mit Sterillium beim Betreten von Hochrisiko-Bereichen (Linie 1, 4, 7)
- Einweg-Handschuhe sind alle 2 Stunden oder bei Beschädigung zu wechseln

2.3 Gesundheit
- Krankmeldepflicht bei Magen-Darm-Beschwerden, Hauterkrankungen, Wunden
- Belehrung nach §43 Infektionsschutzgesetz alle 24 Monate

3. REINIGUNGSPLAN
3.1 Tägliche Reinigung (Ende jeder Schicht)
- Arbeitsflächen: Schaumreiniger Topax 686, Einwirkzeit 10 min, danach Spülen
- Maschinen Linie 3: CIP-Spülung mit alkalischem Reiniger Topax 56 (1,5 %)
- Bodenreinigung: Bona Bodenreiniger 0,5 % alle 4 Stunden
- Wasserschläuche desinfizieren mit Topax 99 nach Gebrauch

3.2 Wöchentliche Tiefenreinigung (jeden Sonntag)
- Komplette Demontage der Schneidemaschinen
- Desinfektion aller Kontaktflächen mit Topax DES400 (3 %)
- Lufthygiene: Filterwechsel und UV-Lampen-Kontrolle
- Abflussreinigung mit Schaumreiniger und Hochdruckspülung

3.3 Monatlich
- Schädlingsbefall-Kontrolle durch externe Firma Anticimex (jeden 1. Montag)
- Wischproben durch QS-Labor: Listeria, E.coli, Gesamtkeimzahl
- Wasserprobe Trinkwasserleitung — extern bei Eurofins

4. DESINFEKTIONSMITTEL UND DOSIERUNG
- Topax 56 (alkalisch): Standardreinigung Maschinen, 1,5-2,0 %
- Topax 99 (Säure): Entkalkung, 0,8-1,2 %, Einwirkzeit 15 min
- Topax 686 (Schaum): Flächen, 3-5 %, Einwirkzeit 10 min
- Topax DES400: Endsäulenfertige Desinfektion, 1,5-3 %
- Sterillium: Händedesinfektion, unverdünnt
WICHTIG: Reiniger niemals mischen — chemische Reaktion möglich.

5. RÜCKSTANDSKONTROLLE
- ATP-Messung an 12 definierten Schwerpunkten nach jeder Reinigung
- Grenzwert: < 30 RLU für Lebensmittelkontaktflächen
- Bei Überschreitung: Nachreinigung und Wiederholungsmessung

6. VERANTWORTLICHKEITEN
- Reinigung Linie 1-3: Schichtführer Produktion (laut Schichtplan)
- QS-Verifikation: Tägliche Stichprobe durch QS-Mitarbeiter:innen
- Gesamtverantwortung: QS-Leitung Karin Petersen
- Externe Firma Anticimex: Schädlingsmonitoring, Berichte monatlich
""",
        encoding="utf-8",
    )

    (popp_dir / "allergeninformation_brotaufstriche.txt").write_text(
        """ALLERGENINFORMATION BROTAUFSTRICHE
Popp Feinkost — Linie 7 (Brotaufstriche)
Dokumenten-Nr.: ALL-BRO-2026-02
Freigegeben durch: QS-Leitung & Produktentwicklung
Letzte Revision: 2026-03-04

1. RECHTLICHE GRUNDLAGE
Diese Information basiert auf der EU-Verordnung 1169/2011 (LMIV), Anhang II,
welche 14 deklarationspflichtige Allergene definiert. Alle Angaben gelten
für Standardrezepturen — Sonderchargen siehe individuelle Spezifikation.

2. PRODUKTÜBERSICHT MIT HAUPTALLERGENEN

2.1 Heringscreme klassisch (Art-Nr. 4501)
Enthält: Fisch (Hering), Senf, Ei (Mayonnaise), Milch (Joghurt), Sulfite
Spuren möglich von: Sellerie, Sesam (gleiche Linie)

2.2 Eiersalat mit Schnittlauch (Art-Nr. 4520)
Enthält: Ei, Senf, Sulfite
Frei von: Fisch, Krebstiere, Erdnüsse, Schalenfrüchte, Sesam, Lupinen
Spuren möglich von: Milch, Sellerie

2.3 Tomaten-Mozzarella-Aufstrich (Art-Nr. 4535)
Enthält: Milch (Mozzarella, Sahne), Ei, Sulfite
Frei von: Fisch, glutenhaltiges Getreide, Erdnüsse

2.4 Champignon-Pfeffer-Creme (Art-Nr. 4540)
Enthält: Milch, Ei, Sellerie, Senf
Spuren möglich von: Sulfite, Soja

2.5 Geflügelsalat fein (Art-Nr. 4550)
Enthält: Ei, Senf, Sellerie, Sulfite
Spuren möglich von: Milch, Sesam

3. KREUZKONTAMINATIONS-MATRIX
Linie 7 verarbeitet auf gleicher Anlage:
- Fischhaltige Produkte (Vor- und Nachmittag)
- Senfhaltige Produkte (alle Schichten)
- Eihaltige Produkte (alle Schichten)

Trennung:
- Fischverarbeitung erfolgt zwischen 06:00-13:00 Uhr
- Vollständige Reinigung gemäß HYG-PRO-2026-04 vor Wechsel
- Allergen-Quick-Test (PCR) nach jeder Wechsel-Reinigung
- Negativbescheinigung erforderlich vor Linienfreigabe

4. UMGANG MIT ANFRAGEN VON KONSUMENTEN
Bei Anfragen zu Spurenangaben:
- Verweis auf gedruckte Packungsdeklaration als verbindlich
- Detailfragen an QS-Leitung weiterleiten (allergene@popp-feinkost.de)
- Bei Zwischenfällen mit Reaktion: sofortige Meldung an QS und Geschäftsführung

5. INTERNE SCHULUNG
- Allergen-Schulung verpflichtend für alle Mitarbeitenden Linie 7 (jährlich)
- Nachweis im Schulungsregister R-SCH-2026
- Verantwortlich: QS Karin Petersen
""",
        encoding="utf-8",
    )


def _seed_popp_machine_text(popp_dir: Path):
    _write_demo_doc(
        popp_dir / "bedienungsanleitung_abfuellanlage_af2400.txt",
        """BEDIENUNGSANLEITUNG ABFÜLLANLAGE AF-2400
Popp Feinkost — Linie 3 Salatabfüllung
Dokumenten-Nr.: BA-AF2400-2026-03
Hersteller: Krones FoodTec GmbH
Anlage: AF-2400, Becherformate 250 g und 500 g
Freigegeben durch: Technik/Instandhaltung
Letzte Revision: 2026-04-18

1. ZWECK UND GELTUNGSBEREICH
Diese Anleitung beschreibt das sichere Starten, Rüsten, Bedienen und Stoppen
der Abfüllanlage AF-2400 auf Linie 3. Sie gilt für Heringssalat, Eiersalat,
Krautsalat und Brotaufstriche mit viskoser Konsistenz.

2. VORAUSSETZUNGEN VOR DEM START
- Linienfreigabe durch QS muss im Terminal L3-QS grün angezeigt werden.
- CIP-Freigabe darf nicht älter als 12 Stunden sein.
- Druckluft muss 6,0 bis 6,5 bar betragen.
- Produktpuffer im Mischer M-7 muss unter +6 °C liegen.
- Bechermagazin, Siegelfolie und Etiketten müssen zum Produktionsauftrag passen.
- Not-Aus-Taster, Schutzhauben und Lichtgitter prüfen.
- Metalldetektor MD-3 muss den Testkörper 0,8 mm Fe erkennen.

3. STARTABLAUF
1. Hauptschalter Q1 auf EIN stellen.
2. HMI mit Bedienerkarte anmelden.
3. Rezeptur über SAP-Auftrag laden, z. B. HSA-250 oder EIS-500.
4. Becherformat bestätigen und Formatwerkzeug kontrollieren.
5. Pumpe P-3 auf Handbetrieb stellen und Produkt bis zum Füllkopf fördern.
6. Vorlaufbecher verwerfen, bis das Produkt blasenfrei austritt.
7. Automatikbetrieb starten und die ersten 10 Becher wiegen.
8. Produktionsfreigabe im Linienprotokoll bestätigen.

4. SOLLWERTE
- Füllgewicht 250-g-Becher: 250 g ± 2 g.
- Füllgewicht 500-g-Becher: 500 g ± 3 g.
- Taktleistung Standard: 42 Becher/min bei 250 g, 30 Becher/min bei 500 g.
- Siegeltemperatur: 178 bis 184 °C.
- Siegeldruck: 4,2 bis 4,8 bar.
- Vakuum Folienansaugung: mindestens -0,55 bar.

5. RÜSTEN AUF ANDERES FORMAT
- Anlage stoppen, Restprodukt über Rücklauf in M-7 führen.
- Becherstern, Füllrohr und Folienführung gemäß Werkzeugliste wechseln.
- Formatwechsel darf nur durch geschulte Operatoren oder Technik erfolgen.
- Nach jedem Rüstvorgang sind 5 Leerläufe und 10 Kontrollbecher Pflicht.
- QS prüft Gewicht, Siegelnaht, Etikett und Metalldetektor.

6. STOPP UND SCHICHTENDE
- Produktzufuhr schließen und Pumpe P-3 leerfahren.
- Anlage in Grundstellung fahren.
- Restbecher aus Austrag entfernen.
- Reinigungsmodus aktivieren und Übergabe an Reinigungsteam dokumentieren.

7. SICHERHEIT
- Schutzhauben niemals überbrücken.
- Bei Eingriffen am Füllkopf muss die Anlage verriegelt und gegen Wiedereinschalten gesichert werden.
- Heiße Siegelbacken nur mit Hitzeschutzhandschuhen berühren.
- Störungen an Servoantrieben nur durch Technik/Instandhaltung beheben lassen.

ANSPRECHPARTNER
Technik Linie 3: Jens Martens — Durchwahl 1715
Produktionsleitung Linie 3: Stefan Brügge — Durchwahl 1432
QS-Freigabe: Karin Petersen — Durchwahl 1810
""",
    )

    _write_demo_doc(
        popp_dir / "stoerungsleitfaden_linie3_abfuellung.txt",
        """STÖRUNGSLEITFADEN LINIE 3 — ABFÜLLUNG UND SIEGELUNG
Popp Feinkost — AF-2400, MD-3, Etikettierer ET-12
Dokumenten-Nr.: STL-L3-2026-02
Freigegeben durch: Technik/Instandhaltung
Letzte Revision: 2026-04-22

1. ALLGEMEINES VORGEHEN
- Störung am HMI lesen und Uhrzeit im Linienprotokoll notieren.
- Produktfluss stoppen, wenn Lebensmittelsicherheit oder Gewicht betroffen ist.
- Bei offener Schutzhaube oder Not-Aus niemals quittieren, bevor die Ursache klar ist.
- Nach jeder produktberührenden Störung QS informieren.
- Wenn eine Störung länger als 15 Minuten dauert: Produktionsleitung und Technik rufen.

2. FEHLER E-17: FÜLLKOPF BLOCKIERT
Symptome:
- HMI meldet E-17 Füllkopf blockiert.
- Füllgewicht schwankt oder Becher bleiben leer.
- Pumpe P-3 läuft gegen Druck.

Maßnahmen:
1. Anlage stoppen und Produktzufuhr schließen.
2. Füllkopf über HMI in Serviceposition fahren.
3. Schutzhaube öffnen und Verriegelung abwarten.
4. Füllrohr auf Produktstücke, Folienreste oder Dichtungsteile prüfen.
5. Füllrohr mit warmem Trinkwasser spülen, keine Metallwerkzeuge verwenden.
6. Dichtung FK-D12 prüfen; bei Beschädigung ersetzen.
7. Anlage schließen, 5 Vorlaufbecher verwerfen und 10 Kontrollbecher wiegen.
8. QS-Freigabe einholen, wenn Produktkontaktflächen geöffnet wurden.

3. FEHLER E-24: SIEGELTEMPERATUR ZU NIEDRIG
Maßnahmen:
- Sollwert 178 bis 184 °C prüfen.
- 8 Minuten Aufheizzeit abwarten.
- Temperaturfühler T-SB1 auf festen Sitz prüfen.
- Wenn Temperatur nicht steigt: Technik rufen, Heizpatrone HZ-2400 prüfen lassen.
- Alle Becher seit letzter gültiger Temperaturmessung sperren.

4. FEHLER E-31: METALLDETEKTOR ABWURF
Maßnahmen:
- Ausgeworfenen Becher sichern und Charge sperren.
- Test mit 0,8 mm Fe, 1,2 mm NFe und 1,5 mm Edelstahl durchführen.
- Bei erfolgreichem Test: betroffenen Becher an QS übergeben.
- Bei fehlgeschlagenem Test: Linie stoppen, Technik und QS rufen.
- Produktion erst nach dokumentierter QS-Freigabe fortsetzen.

5. UNGLEICHMÄSSIGE FÜLLGEWICHTE
Mögliche Ursachen:
- Luft im Produktstrom.
- Produkt zu kalt oder zu fest.
- Füllrohr teilweise blockiert.
- Pumpe P-3 fördert ungleichmäßig.

Korrektur:
- Produktpuffer im Mischer M-7 prüfen, Zieltemperatur +4 bis +6 °C.
- Pumpe P-3 entlüften.
- Rührwerk M-7 auf 6 bis 8 U/min einstellen.
- 20 Kontrollbecher wiegen; bei mehr als 2 Ausreißern Technik rufen.
- Gewichtsabweichungen im SAP-QM Los dokumentieren.

6. ETIKETTENFEHLER ET-12
- Etikettenrolle auf korrekte Artikelnummer prüfen.
- Sensor S-ET2 reinigen.
- Spendeversatz über HMI maximal ±2 mm korrigieren.
- Bei falschem Etikett: Linie stoppen und seit letztem Rollenwechsel Ware sperren.

7. WANN SOFORT TECHNIK RUFEN?
- Servoachse meldet Überstrom.
- Druckluft fällt unter 5,8 bar.
- Siegelbacke heizt über 190 °C.
- Schutzkreis lässt sich nicht quittieren.
- Wiederkehrende E-17 Störung mehr als 2-mal pro Stunde.
""",
    )

    _write_demo_doc(
        popp_dir / "wartungsplan_maschinen_linie3.txt",
        """WARTUNGSPLAN MASCHINEN LINIE 3
Popp Feinkost — Fischsalate und Feinkostsalate
Dokumenten-Nr.: WPL-L3-2026-01
Freigegeben durch: Technik/Instandhaltung
Letzte Revision: 2026-03-30

1. BETROFFENE ANLAGEN
- Treif Puma 700 Schneidemaschine für Heringsfilets.
- Mischer M-7 mit Kühlmantel und langsam laufendem Rührwerk.
- Produktpumpe P-3 und Füllkopf AF-2400.
- Siegelstation SB-2400.
- Metalldetektor MD-3.
- Etikettierer ET-12.
- Kontrollwaage KW-3.

2. TÄGLICHE WARTUNG DURCH PRODUKTION
Vor Produktionsbeginn:
- Sichtprüfung auf Beschädigungen, lose Kabel und Leckagen.
- Druckluft 6,0 bis 6,5 bar dokumentieren.
- Testlauf ohne Produkt für 2 Minuten.
- Not-Aus und Schutzhaubenfunktion prüfen.

Nach Schichtende:
- Produktreste entfernen.
- Füllkopf und Dichtungen auf Risse prüfen.
- Schneidgatter Treif Puma 700 auf Fischgräten und Fremdkörper kontrollieren.
- Kontrollwaage KW-3 reinigen und Nullpunkt prüfen.
- Auffälligkeiten im Linienbuch L3 eintragen.

3. WÖCHENTLICHE WARTUNG DURCH TECHNIK
- Schmierung der Führungsschienen AF-2400 mit NSF-H1-Fett Klüberfood NH1 94-402.
- Prüfung der Dichtung FK-D12 am Füllkopf.
- Kontrolle der Pumpenmembran P-3-MB.
- Prüfung Siegelbacken auf Beschädigung und gleichmäßige Temperatur.
- Test Metalldetektor MD-3 mit Fe, NFe und Edelstahl.
- Etikettierer ET-12: Rollenführung, Sensor S-ET2 und Andruckrolle reinigen.

4. MONATLICHE WARTUNG
- Kalibrierung Kontrollwaage KW-3 mit 100 g, 250 g und 500 g Prüfgewichten.
- Thermofühler T-SB1 und T-M7 mit Referenzthermometer vergleichen.
- Datalogger Kühlmantel M-7 auslesen.
- Sicherheitskreis durch Elektrofachkraft prüfen.
- Ersatzteilbestand kontrollieren.

5. ERSATZTEILE MINDESTBESTAND
- FK-D12 Dichtung Füllkopf: Mindestbestand 12 Stück.
- P-3-MB Pumpenmembran: Mindestbestand 4 Stück.
- HZ-2400 Heizpatrone Siegelbacke: Mindestbestand 2 Stück.
- S-ET2 Etikettensensor: Mindestbestand 1 Stück.
- SB-TF1 Temperaturfühler Siegelbacke: Mindestbestand 2 Stück.
- Messersegment Treif Puma 700 12 mm: Mindestbestand 1 Satz.

6. FREIGABE NACH WARTUNG
- Jede Wartung wird in SAP-PM mit Auftrag und Technikername dokumentiert.
- Nach produktberührender Wartung ist Reinigung und QS-Freigabe Pflicht.
- Nach Eingriff am Metalldetektor oder der Kontrollwaage ist eine Funktionsprüfung Pflicht.
- Ohne dokumentierte Freigabe darf Linie 3 nicht in Produktion gehen.
""",
    )

    _write_demo_doc(
        popp_dir / "cip_reinigung_mischer_m7.txt",
        """CIP-REINIGUNG MISCHER M-7 UND PRODUKTLEITUNG P-3
Popp Feinkost — Linie 3
Dokumenten-Nr.: CIP-M7-2026-04
Freigegeben durch: QS und Technik
Letzte Revision: 2026-04-10

1. ZIEL
Die CIP-Reinigung entfernt Produktreste, Fett, Eiweiß und mikrobiologische
Belastungen aus Mischer M-7, Pumpe P-3, Produktleitung und Füllkopf AF-2400.

2. AUSLÖSER
- Nach jeder Produktionsschicht.
- Bei Produktwechsel von Fisch auf nicht-fischhaltige Artikel.
- Nach ungeplanter Stillstandszeit über 45 Minuten mit Produkt im System.
- Nach Eingriffen an Füllkopf, Pumpe oder Produktleitung.

3. PROGRAMM CIP-L3-STD
1. Vorspülen mit Trinkwasser bei 35 bis 40 °C für 6 Minuten.
2. Alkalische Reinigung Topax 56 mit 1,8 % bei 62 °C für 18 Minuten.
3. Zwischenspülen mit Trinkwasser bis Leitfähigkeit unter 80 µS/cm über Zulauf liegt.
4. Saure Reinigung Topax 99 mit 1,0 % bei 55 °C für 10 Minuten.
5. Klarspülen mit Trinkwasser für 8 Minuten.
6. Desinfektion Topax DES400 mit 1,5 % für 12 Minuten.
7. Endspülen mit kaltem Trinkwasser bis pH 6,5 bis 7,5 erreicht ist.

4. FREIGABEWERTE
- ATP-Messung Lebensmittelkontaktflächen: unter 30 RLU.
- pH-Endspülwasser: 6,5 bis 7,5.
- Leitfähigkeit Endspülwasser: maximal 80 µS/cm über Trinkwasserreferenz.
- Sichtkontrolle: keine Produktreste im Füllkopf oder an Dichtungen.
- Temperaturaufzeichnung muss vollständig im CIP-Report stehen.

5. MANUELLE NACHREINIGUNG
Manuelle Nachreinigung ist erforderlich, wenn:
- Heringsstücke im Füllrohr sichtbar sind.
- ATP-Wert über 30 RLU liegt.
- Dichtung FK-D12 Produktreste zeigt.
- CIP-Programm wegen Druckabfall abgebrochen wurde.

Vorgehen:
- Anlage verriegeln.
- Füllkopf demontieren.
- Kontaktflächen mit Topax 686 einschäumen, 10 Minuten einwirken lassen.
- Mit Trinkwasser spülen und danach Topax DES400 anwenden.
- QS führt erneute ATP-Messung durch.

6. DOKUMENTATION
- CIP-Report wird automatisch unter SAP-QM CIP-L3 gespeichert.
- QS dokumentiert ATP-Werte im Prüfplan QP-L3-CIP.
- Abweichungen werden als NCR angelegt.
- Verantwortlich für Durchführung: Schichtführer Produktion.
- Verantwortlich für Freigabe: QS-Mitarbeiter:in der jeweiligen Schicht.
""",
    )


def _seed_popp_excel(popp_dir: Path):
    # HACCP-Plan
    wb = Workbook()
    ws = wb.active
    ws.title = "HACCP-CCPs"
    ws.append(["Prozessschritt", "Gefahr", "Art", "Krit. Grenzwert", "Überwachung", "Frequenz", "Korrekturmaßnahme", "Verantwortlich"])
    ws.append(["Wareneingang Hering", "Listeria, Histamin", "biologisch", "Histamin < 100 mg/kg, Kerntemperatur < +4 °C", "Temperaturmessung + Stichprobe", "jede Lieferung", "Lieferung sperren, Rücksendung", "QS-Wareneingang"])
    ws.append(["Wässerung Hering", "Salzgehalt zu hoch/niedrig", "chemisch", "Salzgehalt 2,5-3,1 %", "Refraktometer", "alle 30 min", "Wässerungszeit anpassen", "Schichtführer Linie 3"])
    ws.append(["Dressing-Mischung", "Pathogenwachstum", "biologisch", "Produkttemperatur < +6 °C", "Sondenmessung Mischer", "kontinuierlich", "Mischen stoppen, kühlen", "Operator M-7"])
    ws.append(["Abfüllung", "Fremdkörper Metall", "physikalisch", "0 Detektionen über 0,8 mm Fe", "Inline-Metalldetektor", "100 %", "Becher aussortieren, Wartung", "Linientechniker"])
    ws.append(["Siegelung Becher", "Undichtigkeit", "physikalisch", "Drucktest 0,3 bar / 5 s ohne Abfall", "Inline-Vision + Stichprobe", "100 % + 1×/h manuell", "Charge sperren", "QS-Linie"])
    ws.append(["Versand-Lager", "Kühlkette unterbrochen", "biologisch", "Lagertemperatur +2 bis +5 °C", "Datalogger", "kontinuierlich", "Charge prüfen, ggf. sperren", "Lagerleitung"])

    ws2 = wb.create_sheet("Verifikation")
    ws2.append(["CCP", "Verifikationsmethode", "Frequenz", "Verantwortlich", "Doku in"])
    ws2.append(["CCP-1 Wässerung", "Laborprobe Salzgehalt", "wöchentlich", "QS-Labor", "QS-Bericht"])
    ws2.append(["CCP-2 Dressing", "Kalibrierung Sonden", "monatlich", "Externe Firma DKD", "Kalibrierprotokoll"])
    ws2.append(["CCP-3 Abfüllung", "Test-Bumper-Becher mit Metallstücken", "wöchentlich", "Linientechniker", "Logbuch Linie 3"])
    ws2.append(["CCP-4 Siegelung", "Drucktest Stichprobe Labor", "täglich", "QS-Mitarbeiter:in", "QS-Bericht"])
    ws2.append(["CCP-5 Lager", "Datalogger-Auswertung", "wöchentlich", "Lagerleitung", "Lagerbericht"])

    ws3 = wb.create_sheet("Dokumentenlenkung")
    ws3.append(["Feld", "Wert"])
    ws3.append(["Dokument", "HACCP-Plan Popp Feinkost"])
    ws3.append(["Version", "2026-01"])
    ws3.append(["Erstellt", "2026-01-15"])
    ws3.append(["Freigegeben", "2026-01-22"])
    ws3.append(["Nächste Revision", "2027-01-22"])
    ws3.append(["Verantwortlich", "Karin Petersen, QS-Leitung"])
    wb.save(popp_dir / "haccp_plan_2026.xlsx")

    # Rezepturen
    wb = Workbook()
    ws = wb.active
    ws.title = "Heringssalat"
    ws.append(["Zutat", "Menge kg", "Anteil %", "Lieferant", "Artikelnr.", "Allergene"])
    ws.append(["Matjes-Heringsfilets", 38.0, 38.0, "Lemvig Fisheries A/S", "LF-2207", "Fisch"])
    ws.append(["Rote Bete Würfel", 16.0, 16.0, "Nordmark Gemüse GmbH", "NG-118", "-"])
    ws.append(["Mayonnaise 80%", 15.0, 15.0, "Eigenherstellung Linie 5", "EH-MA80", "Ei, Senf"])
    ws.append(["Joghurt 3,5%", 10.0, 10.0, "Molkerei Holstein", "MH-J35", "Milch"])
    ws.append(["Gewürzgurken", 10.0, 10.0, "Eigenherstellung", "EH-GG", "Sulfite"])
    ws.append(["Äpfel Boskoop", 8.0, 8.0, "Obstbau Altes Land", "OAL-BOS", "-"])
    ws.append(["Zwiebeln", 3.0, 3.0, "Nordmark Gemüse GmbH", "NG-202", "-"])
    ws.append(["Summe", 100.0, 100.0, "", "", ""])

    ws2 = wb.create_sheet("Eiersalat")
    ws2.append(["Zutat", "Menge kg", "Anteil %", "Lieferant", "Artikelnr.", "Allergene"])
    ws2.append(["Eier gekocht gewürfelt", 55.0, 55.0, "Eierhof Niedersachsen", "EHN-K12", "Ei"])
    ws2.append(["Mayonnaise 80%", 28.0, 28.0, "Eigenherstellung Linie 5", "EH-MA80", "Ei, Senf"])
    ws2.append(["Joghurt 3,5%", 8.0, 8.0, "Molkerei Holstein", "MH-J35", "Milch"])
    ws2.append(["Schnittlauch frisch", 2.5, 2.5, "Kräuterhof Bremen", "KHB-SL", "-"])
    ws2.append(["Senf mittelscharf", 2.0, 2.0, "Lübecker Senfmühle", "LSM-22", "Senf, Sulfite"])
    ws2.append(["Salz, Pfeffer, Gewürze", 1.5, 1.5, "diverse", "GW-MIX", "-"])
    ws2.append(["Zitronensaft", 3.0, 3.0, "Eigenherstellung", "EH-ZS", "Sulfite"])
    ws2.append(["Summe", 100.0, 100.0, "", "", ""])

    ws3 = wb.create_sheet("Krautsalat")
    ws3.append(["Zutat", "Menge kg", "Anteil %", "Lieferant", "Artikelnr.", "Allergene"])
    ws3.append(["Weißkohl gehobelt", 62.0, 62.0, "Nordmark Gemüse GmbH", "NG-WK1", "-"])
    ws3.append(["Karotten gehobelt", 12.0, 12.0, "Nordmark Gemüse GmbH", "NG-KA2", "-"])
    ws3.append(["Mayonnaise 80%", 14.0, 14.0, "Eigenherstellung Linie 5", "EH-MA80", "Ei, Senf"])
    ws3.append(["Essig 5%", 4.0, 4.0, "Hela Gewürze", "HG-E5", "Sulfite"])
    ws3.append(["Zucker", 3.0, 3.0, "Nordzucker", "NZ-100", "-"])
    ws3.append(["Salz, Kümmel", 1.5, 1.5, "diverse", "GW-KM", "-"])
    ws3.append(["Wasser", 3.5, 3.5, "Trinkwasser", "-", "-"])
    ws3.append(["Summe", 100.0, 100.0, "", "", ""])
    wb.save(popp_dir / "rezepturen.xlsx")

    # Rohstoffe + Lieferanten
    wb = Workbook()
    ws = wb.active
    ws.title = "Lieferanten"
    ws.append(["Lieferant", "Artikel", "Artikelnr.", "MOQ", "Lieferzeit Tage", "Preis EUR/kg", "Zertifizierung", "Ansprechpartner"])
    ws.append(["Lemvig Fisheries A/S", "Matjes-Heringsfilets", "LF-2207", 500, 5, 6.80, "MSC, IFS Food", "Lars Henriksen"])
    ws.append(["Nordmark Gemüse GmbH", "Rote Bete Würfel", "NG-118", 200, 2, 1.20, "QS, Bio (optional)", "Tanja Voss"])
    ws.append(["Nordmark Gemüse GmbH", "Weißkohl gehobelt", "NG-WK1", 300, 1, 0.85, "QS", "Tanja Voss"])
    ws.append(["Nordmark Gemüse GmbH", "Karotten gehobelt", "NG-KA2", 200, 1, 0.95, "QS", "Tanja Voss"])
    ws.append(["Nordmark Gemüse GmbH", "Zwiebeln gewürfelt", "NG-202", 100, 1, 1.10, "QS", "Tanja Voss"])
    ws.append(["Molkerei Holstein", "Joghurt 3,5%", "MH-J35", 1000, 2, 0.95, "IFS Food", "Bernd Albrecht"])
    ws.append(["Eierhof Niedersachsen", "Eier gekocht gewürfelt", "EHN-K12", 200, 3, 4.50, "KAT, IFS", "Birgit Meier"])
    ws.append(["Lübecker Senfmühle", "Senf mittelscharf", "LSM-22", 50, 5, 2.30, "IFS", "Klaus Hansen"])
    ws.append(["Hela Gewürze", "Essig 5%", "HG-E5", 100, 4, 0.65, "IFS", "Anja Schulz"])
    ws.append(["Nordzucker", "Zucker raffiniert", "NZ-100", 1000, 7, 0.78, "IFS, RSPO", "Account Mgmt"])
    ws.append(["Obstbau Altes Land", "Äpfel Boskoop", "OAL-BOS", 500, 3, 1.80, "QS, GlobalGAP", "Ernst Brandt"])
    ws.append(["Kräuterhof Bremen", "Schnittlauch frisch", "KHB-SL", 20, 2, 12.50, "QS", "Frieda Kowalski"])

    ws2 = wb.create_sheet("Allergene-Mapping")
    ws2.append(["Rohstoff", "Allergen", "Hinweis Verarbeitung"])
    ws2.append(["Matjes-Heringsfilets", "Fisch", "Linie 1-3, separater Bereich"])
    ws2.append(["Mayonnaise 80%", "Ei, Senf", "Eigenherstellung Linie 5"])
    ws2.append(["Joghurt 3,5%", "Milch", "Lagerung getrennt von Fisch"])
    ws2.append(["Eier gekocht", "Ei", "Linie 7 (Brotaufstriche/Salate)"])
    ws2.append(["Senf mittelscharf", "Senf, Sulfite", "Standard"])
    ws2.append(["Essig 5%", "Sulfite", "Standard"])
    ws2.append(["Gewürzgurken", "Sulfite", "Eigenherstellung"])

    ws3 = wb.create_sheet("Mindestbestand")
    ws3.append(["Artikelnr.", "Artikel", "Mindestbestand kg", "Aktueller Bestand kg", "Status", "Nachbestellung bis"])
    ws3.append(["LF-2207", "Matjes-Heringsfilets", 800, 1240, "OK", ""])
    ws3.append(["NG-118", "Rote Bete Würfel", 300, 280, "Niedrig", "2026-05-08"])
    ws3.append(["MH-J35", "Joghurt 3,5%", 500, 720, "OK", ""])
    ws3.append(["EHN-K12", "Eier gekocht gewürfelt", 200, 95, "Kritisch", "2026-05-07"])
    ws3.append(["LSM-22", "Senf mittelscharf", 80, 110, "OK", ""])
    ws3.append(["NG-WK1", "Weißkohl gehobelt", 400, 480, "OK", ""])
    wb.save(popp_dir / "rohstoffe_lieferanten.xlsx")


# ──────────────────────────────────────────────
# Lifespan
# ──────────────────────────────────────────────
@asynccontextmanager
async def lifespan(_app: FastAPI):
    _migrate_legacy_docs()
    _seed_popp_demo()
    for ws_id in WORKSPACES:
        sync_docs_folder(ws_id)
    yield


# ──────────────────────────────────────────────
# App
# ──────────────────────────────────────────────
app = FastAPI(title="Noxera Labs AI Chat", version=APP_VERSION, lifespan=lifespan)

cors_origins = env_list("CORS_ORIGINS", "*")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────────────────────────────
# Models
# ──────────────────────────────────────────────
class QueryRequest(BaseModel):
    question: str
    n_results: int = 5
    model: Optional[str] = None
    web_search: bool = False
    history: list[dict] = []
    workspace: str = DEFAULT_WORKSPACE


# ──────────────────────────────────────────────
# Retrieval
# ──────────────────────────────────────────────
def retrieve(ws_id: str, question: str, n_results: int) -> tuple[list[str], list[dict], list[float]]:
    coll = collection_for(ws_id)
    total = coll.count()
    if total == 0:
        return [], [], []
    n = min(n_results, total)
    results = coll.query(query_texts=[question], n_results=n)
    chunks = results.get("documents", [[]])[0] or []
    metas = results.get("metadatas", [[]])[0] or []
    dists = results.get("distances", [[]])[0] or []

    seen: set = set()
    out_chunks, out_metas, out_dists = [], [], []
    for chunk, meta in lexical_matches(coll, question):
        key = (meta.get("filename"), meta.get("sheet_name"), meta.get("chunk_index"))
        if key in seen:
            continue
        seen.add(key)
        out_chunks.append(chunk)
        out_metas.append(meta)
        out_dists.append(0.0)
        if len(out_chunks) >= n:
            return out_chunks, out_metas, out_dists

    for chunk, m, d in zip(chunks, metas, dists):
        relevance = 1 - d
        if relevance < RAG_MIN_RELEVANCE:
            continue
        key = (m.get("filename"), m.get("sheet_name"), m.get("chunk_index"))
        if key in seen:
            continue
        seen.add(key)
        out_chunks.append(chunk)
        out_metas.append(m)
        out_dists.append(d)
        if len(out_chunks) >= n:
            break
    return out_chunks, out_metas, out_dists


def lexical_matches(coll, question: str) -> list[tuple[str, dict]]:
    q = question.lower()
    terms = [t for t in re.findall(r"[a-zäöüß0-9]{4,}", q) if t not in {"welche", "exakte"}]
    try:
        all_items = coll.get()
    except Exception:
        return []

    scored: list[tuple[int, str, dict]] = []
    for chunk, meta in zip(all_items.get("documents", []) or [], all_items.get("metadatas", []) or []):
        if not chunk or not meta:
            continue
        text = chunk.lower()
        filename = (meta.get("filename") or "").lower()
        sheet_name = (meta.get("sheet_name") or "").lower()
        score = 0
        if sheet_name and sheet_name in q:
            score += 4
        if filename.endswith(".xlsx") and ("rezeptur" in q or "excel" in q or "tabelle" in q):
            score += 2
        score += min(4, sum(1 for term in terms if term in text))
        if score >= 5:
            scored.append((score, chunk, meta))

    scored.sort(key=lambda item: item[0], reverse=True)
    return [(chunk, meta) for _, chunk, meta in scored[:3]]


def build_messages(question: str, chunks: list[str], metadatas: list[dict], history: Optional[list[dict]] = None) -> list[dict]:
    messages: list[dict] = []
    for msg in (history or []):
        role = "assistant" if msg.get("role") == "bot" else "user"
        content = str(msg.get("content", "")).strip()
        if not content:
            continue
        if messages and messages[-1]["role"] == role:
            continue
        messages.append({"role": role, "content": content})
    if messages and messages[-1]["role"] == "user":
        messages.pop()

    if not chunks:
        final = question
    else:
        context = "\n\n---\n\n".join(
            f"[{i+1}] Quelle: {source_label(m)}, Abschnitt {m.get('chunk_index', 0)+1}\n{chunk}"
            for i, (chunk, m) in enumerate(zip(chunks, metadatas))
        )
        final = f"Kontext:\n{context}\n\nFrage: {question}"
    messages.append({"role": "user", "content": final})
    return messages


def source_label(meta: dict) -> str:
    filename = meta.get("filename", "Unbekannt")
    sheet_name = (meta.get("sheet_name") or "").strip()
    if sheet_name:
        return f"{filename} / {sheet_name}"
    return filename


def no_context_answer(ws_id: str) -> Optional[str]:
    if ws_id != "popp":
        return None
    return (
        "Diese Information ist nicht in der freigegebenen Wissensbasis enthalten. "
        "Bitte die QS-Leitung oder die zuständige Fachabteilung kontaktieren."
    )


def sources_payload(metadatas: list[dict], distances: list[float]) -> list[dict]:
    return [
        {
            "index": i + 1,
            "filename": source_label(m),
            "chunk_index": (m.get("chunk_index") or 0) + 1,
            "relevance_score": round(1 - dist, 4),
        }
        for i, (m, dist) in enumerate(zip(metadatas, distances))
    ]


def sse(event: str, data: dict) -> bytes:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n".encode()


def text_from_response(resp) -> str:
    parts = []
    for block in getattr(resp, "content", []) or []:
        if getattr(block, "type", "") == "text":
            parts.append(getattr(block, "text", ""))
    return "".join(parts)


def request_error(req: QueryRequest) -> Optional[str]:
    if not req.question.strip():
        return "Bitte gib eine Frage ein."
    if len(req.question) > MAX_QUESTION_CHARS:
        return f"Die Frage ist zu lang. Limit: {MAX_QUESTION_CHARS} Zeichen."
    if req.model and req.model not in ALLOWED_MODELS:
        return "Dieses Modell ist für diese Demo nicht freigegeben."
    return None


def model_for(req: QueryRequest) -> str:
    return req.model or DEFAULT_MODEL


async def extract_text_from_image(content: bytes, media_type: str) -> str:
    b64 = base64.b64encode(content).decode("ascii")
    response = await anthropic_client.messages.create(
        model=OCR_MODEL,
        max_tokens=1200,
        system=(
            "Du extrahierst Text aus Dokumentenfotos für ein RAG-System. "
            "Gib nur den erkannten Text zurück. Erfinde keine Inhalte. "
            "Wenn kaum Text erkennbar ist, schreibe: KEIN_TEXT_ERKANNT."
        ),
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": "Extrahiere den sichtbaren Text aus diesem Bild möglichst vollständig.",
                    },
                ],
            }
        ],
    )
    return text_from_response(response).strip()


# ──────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────
@app.get("/workspaces")
async def list_workspaces():
    return {
        "workspaces": [
            {
                "id": w["id"],
                "name": w["name"],
                "tagline": w["tagline"],
                "description": w["description"],
                "accent": w["accent"],
                "logo_url": w.get("logo_url"),
                "allow_web_search": w.get("allow_web_search", False),
                "chips": w["chips"],
            }
            for w in WORKSPACES.values()
        ],
        "default": DEFAULT_WORKSPACE,
    }


@app.post("/query/stream")
async def query_stream(req: QueryRequest):
    err = request_error(req)
    ws_id = workspace_id_or_default(req.workspace)
    model = model_for(req)
    n_results = max(1, min(req.n_results, MAX_RETRIEVAL_RESULTS))
    chunks, metadatas, distances = retrieve(ws_id, req.question, n_results) if not err else ([], [], [])

    async def event_generator() -> AsyncIterator[bytes]:
        if err:
            yield sse("error", {"message": err})
            return
        if chunks:
            yield sse("sources", {"sources": sources_payload(metadatas, distances)})
        no_context = no_context_answer(ws_id)
        if not chunks and no_context:
            yield sse("token", {"text": no_context})
            yield sse("done", {})
            return

        try:
            stream_kwargs: dict = dict(
                model=model,
                max_tokens=MAX_RESPONSE_TOKENS,
                system=system_prompt_for(ws_id),
                messages=build_messages(req.question, chunks, metadatas, req.history),
            )
            if req.web_search and web_search_allowed(ws_id):
                stream_kwargs["tools"] = [{"type": WEB_SEARCH_TOOL_TYPE, "name": "web_search"}]
                yield sse("status", {"text": "Web-Suche läuft…"})

            async with anthropic_client.messages.stream(**stream_kwargs) as stream:
                async for text in stream.text_stream:
                    yield sse("token", {"text": text})
            yield sse("done", {})
        except Exception as e:
            yield sse("error", {"message": str(e)})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/query")
async def query(req: QueryRequest):
    err = request_error(req)
    if err:
        return JSONResponse({"status": "error", "message": err}, status_code=400)
    ws_id = workspace_id_or_default(req.workspace)
    model = model_for(req)
    n_results = max(1, min(req.n_results, MAX_RETRIEVAL_RESULTS))
    chunks, metadatas, distances = retrieve(ws_id, req.question, n_results)
    no_context = no_context_answer(ws_id)
    if not chunks and no_context:
        return {
            "answer": no_context,
            "sources": [],
            "model": model,
            "workspace": ws_id,
        }
    stream_kwargs: dict = dict(
        model=model,
        max_tokens=MAX_RESPONSE_TOKENS,
        system=system_prompt_for(ws_id),
        messages=build_messages(req.question, chunks, metadatas, req.history),
    )
    if req.web_search and web_search_allowed(ws_id):
        stream_kwargs["tools"] = [{"type": WEB_SEARCH_TOOL_TYPE, "name": "web_search"}]

    try:
        response = await anthropic_client.messages.create(**stream_kwargs)
        return {
            "answer": text_from_response(response),
            "sources": sources_payload(metadatas, distances) if chunks else [],
            "model": model,
            "workspace": ws_id,
        }
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)


@app.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    workspace: str = Query(DEFAULT_WORKSPACE),
):
    ws_id = workspace_id_or_default(workspace)
    filename = safe_filename(file.filename or "upload")
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_DOC_EXTS:
        return JSONResponse(
            {"status": "error", "message": "Nur PDF, TXT und XLSX werden unterstützt"},
            status_code=400,
        )
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        return JSONResponse(
            {"status": "error", "message": "Datei ist zu groß für diese Demo"},
            status_code=413,
        )
    try:
        if suffix == ".pdf":
            text = extract_pdf_text(content)
        elif suffix == ".xlsx":
            text = extract_excel_text(content)
        else:
            text = content.decode("utf-8", errors="ignore").strip()
    except Exception as e:
        return JSONResponse(
            {"status": "error", "message": f"Datei konnte nicht gelesen werden: {e}"},
            status_code=400,
        )

    if not text:
        return {"status": "empty", "chunks": 0, "filename": filename, "workspace": ws_id}

    content_hash = file_hash(content)
    stored_name = f"{Path(filename).stem}--{content_hash}{suffix}"
    stored_path = docs_dir_for(ws_id) / stored_name
    if not stored_path.exists():
        stored_path.write_bytes(content)

    if suffix == ".xlsx":
        status, n, _ = _index_excel_bytes(ws_id, stored_name, content, stored_name)
    else:
        status, n = _index_text(ws_id, stored_name, text, stored_name)
    return {
        "status": status,
        "chunks": n,
        "filename": stored_name,
        "stored": True,
        "workspace": ws_id,
    }


@app.post("/upload-image")
async def upload_image(
    file: UploadFile = File(...),
    workspace: str = Query(DEFAULT_WORKSPACE),
):
    ws_id = workspace_id_or_default(workspace)
    filename = safe_filename(file.filename or "scan.jpg")
    suffix = Path(filename).suffix.lower()
    media_type = IMAGE_MEDIA_TYPES.get(suffix)
    if not media_type:
        return JSONResponse(
            {"status": "error", "message": "Nur JPG, PNG und WEBP werden unterstützt"},
            status_code=400,
        )

    content = await file.read()
    if len(content) > MAX_IMAGE_UPLOAD_BYTES:
        return JSONResponse(
            {"status": "error", "message": "Bild ist zu groß für die OCR-Demo"},
            status_code=413,
        )

    try:
        text = await extract_text_from_image(content, media_type)
    except Exception as e:
        return JSONResponse({"status": "error", "message": f"OCR fehlgeschlagen: {e}"}, status_code=500)

    if not text or text.upper().startswith("KEIN_TEXT_ERKANNT"):
        return {"status": "empty", "chunks": 0, "filename": filename, "workspace": ws_id, "message": "Kein lesbarer Text erkannt"}

    content_hash = file_hash(content)
    stored_name = f"{Path(filename).stem}--scan--{content_hash}.txt"
    stored_path = docs_dir_for(ws_id) / stored_name
    stored_path.write_text(
        f"Quelle: Kamerascan aus {filename}\n\n{text}",
        encoding="utf-8",
    )
    status, n = _index_text(ws_id, stored_name, stored_path.read_text(encoding="utf-8"), stored_name)
    return {
        "status": status,
        "chunks": n,
        "filename": stored_name,
        "stored": True,
        "ocr_model": OCR_MODEL,
        "workspace": ws_id,
    }


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "version": APP_VERSION,
        "model": DEFAULT_MODEL,
        "rag_min_relevance": RAG_MIN_RELEVANCE,
        "max_response_tokens": MAX_RESPONSE_TOKENS,
        "max_retrieval_results": MAX_RETRIEVAL_RESULTS,
        "default_workspace": DEFAULT_WORKSPACE,
        "workspaces": {
            ws_id: {
                "name": w["name"],
                "chunks_indexed": collection_for(ws_id).count(),
                "docs": docs_summary(ws_id),
            }
            for ws_id, w in WORKSPACES.items()
        },
    }


@app.get("/knowledge")
async def knowledge_status(workspace: str = Query(DEFAULT_WORKSPACE)):
    ws_id = workspace_id_or_default(workspace)
    return docs_payload(ws_id)


@app.post("/admin/reindex")
async def reindex(workspace: str = Query(DEFAULT_WORKSPACE)):
    ws_id = workspace_id_or_default(workspace)
    summary = sync_docs_folder(ws_id)
    return {
        "status": "ok",
        "workspace": ws_id,
        "summary": summary,
        "total_chunks": collection_for(ws_id).count(),
    }


# ──────────────────────────────────────────────
# Static (must be last)
# ──────────────────────────────────────────────
if STATIC_DIR.exists():
    app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")
